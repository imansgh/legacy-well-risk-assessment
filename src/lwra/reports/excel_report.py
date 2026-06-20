"""Excel report generation (OpenPyXL).

Writes a multi-sheet ``.xlsx`` workbook for a single well assessment:

* **Inputs** -- the static :class:`~lwra.models.well.WellData`, casing strings,
  and observed barriers;
* **Integrity** -- component and overall integrity scores, category, flags;
* **Risk** -- scalar/likelihood/consequence scores, category, weighted factors,
  dominant drivers, matrix cell;
* **Recommendation** -- verdict, suitability levels, confidence, rationale,
  required actions;
* **Calculation Traces** -- the full combined trace flattened to (key, value)
  rows, present only when the assessment carries a trace.

Pure, deterministic functions returning the written file path. The workbook is
analyst-facing: it carries the numbers and the full audit trail.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from lwra.models.well import WellData
from lwra.reports._common import (
    DISCLAIMER,
    check_consistency,
    flatten,
    fmt,
    resolve_output_path,
)
from lwra.services.pipeline import WellAssessment

__all__ = ["write_excel_report", "DEFAULT_OUTPUT_DIR"]

DEFAULT_OUTPUT_DIR: str = "reports_out"

_HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_TITLE_FONT = Font(bold=True, size=14, color="1F4E79")
_SECTION_FONT = Font(bold=True, size=11, color="1F4E79")
_WRAP = Alignment(wrap_text=True, vertical="top")


def _write_header_row(ws: Worksheet, row: int, headers: list[str]) -> None:
    """Write and style a table header row.

    Args:
        ws: Target worksheet.
        row: 1-based row index to write into.
        headers: Column header labels.

    """
    for col, label in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=label)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT


def _autosize(ws: Worksheet, *, max_width: int = 80) -> None:
    """Approximate column auto-sizing from cell content length.

    Args:
        ws: Worksheet to size.
        max_width: Upper bound on column width in characters.

    """
    widths: dict[int, int] = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            length = max(len(line) for line in str(cell.value).split("\n"))
            widths[cell.column] = max(widths.get(cell.column, 0), length)
    for col, width in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = min(width + 2, max_width)


def _sheet_inputs(ws: Worksheet, well: WellData) -> None:
    """Populate the Inputs sheet with well, casing, and barrier data.

    Args:
        ws: The (already-titled) worksheet.
        well: The source well data.

    """
    ws.cell(row=1, column=1, value=f"Well Inputs \u2014 {well.well_id}").font = _TITLE_FONT

    row = 3
    ws.cell(row=row, column=1, value="Well attributes").font = _SECTION_FONT
    row += 1
    _write_header_row(ws, row, ["Field", "Value"])
    row += 1
    attrs: list[tuple[str, object]] = [
        ("Well ID", well.well_id),
        ("Name", well.name),
        ("Latitude", well.location.latitude),
        ("Longitude", well.location.longitude),
        ("Spud date", well.spud_date),
        ("Abandonment date", well.abandonment_date),
        ("Total depth (m)", well.total_depth_m),
        ("Well type", well.well_type),
        ("Formation", well.formation),
        ("Reservoir fluid", well.reservoir_fluid),
        ("Pressure (bar)", well.pressure_bar),
        ("Temperature (deg C)", well.temperature_c),
        ("Proximity to receptors (m)", well.proximity_to_receptors_m),
        ("Abandoned", well.is_abandoned),
    ]
    for label, value in attrs:
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=fmt(value))
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Casing strings").font = _SECTION_FONT
    row += 1
    casing_headers = ["Name", "OD (in)", "Top (m)", "Bottom (m)", "Cemented"]
    _write_header_row(ws, row, casing_headers)
    row += 1
    if well.casing_strings:
        for s in well.casing_strings:
            ws.cell(row=row, column=1, value=s.name)
            ws.cell(row=row, column=2, value=s.outer_diameter_in)
            ws.cell(row=row, column=3, value=s.depth_top_m)
            ws.cell(row=row, column=4, value=s.depth_bottom_m)
            ws.cell(row=row, column=5, value=fmt(s.cemented))
            row += 1
    else:
        ws.cell(row=row, column=1, value="(none recorded)")
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Barriers").font = _SECTION_FONT
    row += 1
    barrier_headers = [
        "Barrier ID",
        "Type",
        "Element",
        "Top (m)",
        "Bottom (m)",
        "Condition (0-1)",
        "Verified",
        "Verification method",
    ]
    _write_header_row(ws, row, barrier_headers)
    row += 1
    if well.barriers:
        for b in well.barriers:
            ws.cell(row=row, column=1, value=b.barrier_id)
            ws.cell(row=row, column=2, value=fmt(b.barrier_type))
            ws.cell(row=row, column=3, value=fmt(b.element))
            ws.cell(row=row, column=4, value=b.depth_top_m)
            ws.cell(row=row, column=5, value=b.depth_bottom_m)
            ws.cell(row=row, column=6, value=b.condition_score)
            ws.cell(row=row, column=7, value=fmt(b.verified))
            ws.cell(row=row, column=8, value=b.verification_method or "\u2013")
            row += 1
    else:
        ws.cell(row=row, column=1, value="(none recorded)")
        row += 1

    _autosize(ws)


def _sheet_integrity(ws: Worksheet, assessment: WellAssessment) -> None:
    """Populate the Integrity sheet.

    Args:
        ws: The worksheet.
        assessment: The assessment being reported.

    """
    integ = assessment.integrity
    ws.cell(row=1, column=1, value=f"Integrity \u2014 {integ.well_id}").font = _TITLE_FONT

    row = 3
    _write_header_row(ws, row, ["Component", "Score (0-100)"])
    row += 1
    components: list[tuple[str, float]] = [
        ("Primary barrier", integ.primary_barrier_score),
        ("Secondary barrier", integ.secondary_barrier_score),
        ("Cement quality", integ.cement_quality_score),
        ("Mechanical integrity", integ.mechanical_integrity_score),
        ("Plugging", integ.plugging_score),
        ("OVERALL", integ.overall_integrity_score),
    ]
    for label, score in components:
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=score)
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Category")
    ws.cell(row=row, column=2, value=fmt(integ.integrity_category))
    row += 2

    ws.cell(row=row, column=1, value="Flags").font = _SECTION_FONT
    row += 1
    if integ.flags:
        for flag in integ.flags:
            ws.cell(row=row, column=1, value=flag)
            row += 1
    else:
        ws.cell(row=row, column=1, value="(no flags)")
        row += 1

    _autosize(ws)


def _sheet_risk(ws: Worksheet, assessment: WellAssessment) -> None:
    """Populate the Risk sheet.

    Args:
        ws: The worksheet.
        assessment: The assessment being reported.

    """
    risk = assessment.risk
    ws.cell(row=1, column=1, value=f"Risk \u2014 {risk.well_id}").font = _TITLE_FONT

    row = 3
    _write_header_row(ws, row, ["Metric", "Value"])
    row += 1
    metrics: list[tuple[str, object]] = [
        ("Risk score (0-100)", risk.risk_score),
        ("Risk category", fmt(risk.risk_category)),
        ("Likelihood (0-100)", risk.likelihood),
        ("Consequence (0-100)", risk.consequence),
    ]
    coords = (risk.calculation_trace or {}).get("matrix_coordinates", {})
    if coords:
        metrics.append(
            (
                "Matrix cell",
                f"L{coords.get('likelihood_bin')}-C{coords.get('consequence_bin')}",
            )
        )
    for label, value in metrics:
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=value)
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Weighted factor contributions").font = _SECTION_FONT
    row += 1
    _write_header_row(ws, row, ["Factor", "Weighted contribution"])
    row += 1
    for factor, contribution in sorted(
        risk.weighted_factors.items(), key=lambda kv: kv[1], reverse=True
    ):
        ws.cell(row=row, column=1, value=factor)
        ws.cell(row=row, column=2, value=contribution)
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Dominant risk drivers").font = _SECTION_FONT
    row += 1
    if risk.dominant_risk_drivers:
        for driver in risk.dominant_risk_drivers:
            ws.cell(row=row, column=1, value=driver)
            row += 1
    else:
        ws.cell(row=row, column=1, value="(none)")
        row += 1

    _autosize(ws)


def _sheet_recommendation(ws: Worksheet, assessment: WellAssessment) -> None:
    """Populate the Recommendation sheet.

    Args:
        ws: The worksheet.
        assessment: The assessment being reported.

    """
    rec = assessment.recommendation
    ws.cell(row=1, column=1, value=f"Recommendation \u2014 {rec.well_id}").font = _TITLE_FONT

    row = 3
    _write_header_row(ws, row, ["Field", "Value"])
    row += 1
    fields: list[tuple[str, object]] = [
        ("Verdict", fmt(rec.verdict)),
        ("CO2 storage suitability", fmt(rec.co2_storage_suitability)),
        ("Geothermal suitability", fmt(rec.geothermal_suitability)),
        ("Confidence (0-1)", rec.confidence),
    ]
    for label, value in fields:
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=value)
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Required actions").font = _SECTION_FONT
    row += 1
    if rec.required_actions:
        for index, action in enumerate(rec.required_actions, start=1):
            ws.cell(row=row, column=1, value=index)
            ws.cell(row=row, column=2, value=action)
            row += 1
    else:
        ws.cell(row=row, column=1, value="(none required)")
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Rationale").font = _SECTION_FONT
    row += 1
    cell = ws.cell(row=row, column=1, value=rec.rationale)
    cell.alignment = _WRAP
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)

    _autosize(ws)


def _sheet_traces(ws: Worksheet, assessment: WellAssessment) -> None:
    """Populate the Calculation Traces sheet from the flattened combined trace.

    Args:
        ws: The worksheet.
        assessment: The assessment being reported (must carry a trace).

    """
    header_cell = ws.cell(row=1, column=1, value=f"Calculation Traces \u2014 {assessment.well_id}")
    header_cell.font = _TITLE_FONT
    row = 3
    _write_header_row(ws, row, ["Trace key", "Value"])
    row += 1
    flat = flatten(assessment.trace or {})
    for key, value in flat.items():
        ws.cell(row=row, column=1, value=key)
        cell_value = fmt(value) if not isinstance(value, (int, float)) else value
        ws.cell(row=row, column=2, value=cell_value)
        row += 1
    _autosize(ws, max_width=100)


def write_excel_report(
    well: WellData,
    assessment: WellAssessment,
    output_path: str | Path | None = None,
) -> Path:
    """Write a multi-sheet Excel report and return its path.

    Args:
        well: The source well data (for the Inputs sheet).
        assessment: The completed assessment. If it carries a combined trace, a
            ``Calculation Traces`` sheet is added.
        output_path: Destination file or directory, or ``None`` to use a
            generated filename under :data:`DEFAULT_OUTPUT_DIR`.

    Returns:
        The path to the written ``.xlsx`` file.

    Raises:
        ValueError: If ``well`` and ``assessment`` refer to different wells.

    """
    check_consistency(well, assessment)

    wb = Workbook()
    wb.properties.title = f"LWRA Assessment {assessment.well_id}"
    wb.properties.subject = "Legacy Well Risk Assessment"
    wb.properties.description = DISCLAIMER

    inputs_ws = wb.active
    inputs_ws.title = "Inputs"
    _sheet_inputs(inputs_ws, well)
    _sheet_integrity(wb.create_sheet("Integrity"), assessment)
    _sheet_risk(wb.create_sheet("Risk"), assessment)
    _sheet_recommendation(wb.create_sheet("Recommendation"), assessment)
    if assessment.trace is not None:
        _sheet_traces(wb.create_sheet("Calculation Traces"), assessment)

    target = resolve_output_path(
        output_path,
        well_id=assessment.well_id,
        default_dir=DEFAULT_OUTPUT_DIR,
        suffix=".xlsx",
    )
    wb.save(str(target))
    return target
